import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, NamedStyle, fills, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import range_boundaries
from openpyxl.utils.cell import coordinate_from_string

from pyexcel.base import BaseExcelCreator


class OpenPyExcelCreator(BaseExcelCreator):
    side = Side(border_style='thin', color=Color('00000000'))
    header_default_style = NamedStyle(name='header', font=Font(size=14, bold=False, name='Arial'),
                                      fill=PatternFill(patternType=fills.FILL_PATTERN_LIGHTGRAY),
                                      alignment=Alignment(horizontal='center', vertical='center'),
                                      border=Border(bottom=side, left=side, right=side))
    body_default_style = NamedStyle(name='body', font=Font(size=13, bold=False, name='Arial'),
                                    alignment=Alignment(horizontal='center', vertical='center'))

    @staticmethod
    def create_new_workbook(encoding='utf-8') -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        wb.encoding = encoding
        return wb

    @staticmethod
    def add_sheet(workbook: openpyxl.Workbook, sheet_name: str) -> openpyxl.Workbook:
        workbook.create_sheet(title=sheet_name, index=0)
        return workbook

    @classmethod
    def _write_data_in_sheet(cls,
                             workbook: openpyxl.Workbook,
                             sheet_name: str,
                             data: list,
                             header_style: NamedStyle = header_default_style,
                             body_style: NamedStyle = body_default_style,
                             batch_size=1000) -> openpyxl.Workbook:
        """
        :param workbook: object of Workbook
        :param sheet_name: name of sheet which is Written
        :param data: data of sheet
        :param header_style: style of excel header cells
        :param body_style: style of excel body cells except headers
        :param batch_size: size of rows for one writing
        :return: updated Workbook
        """
        worksheet = workbook.get_sheet_by_name(name=sheet_name)

        # compute the range of cells to write
        start_cell = "A1"
        end_cell = coordinate_from_string(start_cell)[0] + str(len(data[0]) - 1)
        range_string = f"{start_cell}:{end_cell}{len(data)}"
        min_col, min_row, max_col, max_row = range_boundaries(range_string)

        # write the data in batches
        for i in range(0, len(data), batch_size):
            rows = data[i:i + batch_size]
            for j, row in enumerate(rows):
                for k, value in enumerate(row):
                    cell = worksheet.cell(row=min_row + i + j, column=min_col + k, value=value)
                    cell.style = body_style

        # apply header style
        for cell in worksheet[1]:
            cell.style = header_style

        for col in worksheet.columns:
            worksheet.column_dimensions[col[0].column_letter].width = 25

        return workbook

    def create_workbook_single_sheet(self,
                                     headers_dict: dict,
                                     input_data: list,
                                     sheet_name=str(datetime.date.today()),
                                     header_style: NamedStyle = header_default_style,
                                     body_style: NamedStyle = body_default_style,
                                     workbook: openpyxl.Workbook = None,
                                     record_type: str = 'json',
                                     encoding='utf-8'):
        """
        :param headers_dict: a dict of headers with english keys and translated values
        :param input_data: list of data records
        :param sheet_name: name of sheet else excel creation date
        :param header_style: style of excel header cells
        :param body_style: style of excel body cells except headers
        :param workbook: exist workbook else it will be created
        :param record_type: define records type: json or list
        :param encoding: default is utf-8
        :return: openpyxl workbook instance
        """

        workbook = workbook or self.create_new_workbook(encoding=encoding)

        workbook = self.add_sheet(workbook=workbook, sheet_name=sheet_name)
        cleaned_data = self.clean_data(headers_dict=headers_dict, input_data=input_data, record_type=record_type)
        workbook = self._write_data_in_sheet(workbook=workbook, sheet_name=sheet_name, data=cleaned_data,
                                             header_style=header_style, body_style=body_style)
        return workbook
