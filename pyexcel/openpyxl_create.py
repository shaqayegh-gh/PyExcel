import datetime
from collections import OrderedDict

import openpyxl
from openpyxl.styles import PatternFill, Font, NamedStyle, fills, Alignment
from openpyxl.utils import range_boundaries
from openpyxl.utils.cell import coordinate_from_string


class ExcelCreator:

    @staticmethod
    def _validate_input_data(input_data: list, record_type: str):
        if record_type == 'json':
            record_type_obj = [dict, OrderedDict]
        elif record_type == 'list':
            record_type_obj = [list]
        else:
            raise Exception("record_type must be json or list")
        if isinstance(input_data, list):
            if input_data and not all([type(item) in record_type_obj for item in input_data]):
                raise TypeError(f'All items of input data should be the same type of {record_type}')
            return input_data
        raise TypeError('Input data should be a list')

    @staticmethod
    def convert_json_data(headers_dict: dict, input_data: list):
        """
        :param headers_dict: a dict of headers with english keys and translated values
        :param input_data: list of data records
        :return: convert json data type to list of data with values
        """
        headers_list = list(headers_dict.keys())
        data = [list({key: item[key] for key in headers_list if key in item}.values()) for item in input_data]
        return data

    @staticmethod
    def create_new_workbook(encoding='utf-8') -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        wb.encoding = encoding
        return wb

    @staticmethod
    def add_sheet(workbook: openpyxl.Workbook, sheet_name: str) -> openpyxl.Workbook:
        workbook.create_sheet(title=sheet_name, index=0)
        return workbook

    def write_data_to_sheet(self, workbook: openpyxl.Workbook, sheet_name: str, data: list) -> openpyxl.Workbook:
        """
        :param workbook: object of Workbook
        :param sheet_name: name of sheet which is Written
        :param data: data of sheet
        :return: updated Workbook
        """
        worksheet = workbook.get_sheet_by_name(name=sheet_name)

        # compute the range of cells to write
        start_cell = "A1"
        end_cell = coordinate_from_string(start_cell)[0] + str(len(data[0]) - 1)
        range_string = f"{start_cell}:{end_cell}{len(data)}"
        min_col, min_row, max_col, max_row = range_boundaries(range_string)

        style = self.get_table_default_style()
        # write the data in batches
        batch_size = 300
        for i in range(0, len(data), batch_size):
            rows = data[i:i + batch_size]
            for j, row in enumerate(rows):
                for k, value in enumerate(row):
                    cell = worksheet.cell(row=min_row + i + j, column=min_col + k, value=value)
                    cell.style = style

        return workbook

    def clean_data(self, headers_dict: dict, input_data: list, record_type: str):
        """
        :param headers_dict: a dict of headers with english keys and translated values
        :param input_data: list of data records
        :param record_type: define records type: json or list
        :return:
        """
        validated_data = self._validate_input_data(input_data=input_data, record_type=record_type)
        if record_type == 'json':
            validated_data = self.convert_json_data(headers_dict=headers_dict, input_data=validated_data)
        # add headers to first index of data list
        validated_data.insert(0, list(headers_dict.values()))
        return validated_data

    @classmethod
    def get_header_default_style(cls):
        header_style = NamedStyle(name='header', font=Font(size=13, bold=True, name='Arial'),
                                  fill=PatternFill(patternType=fills.FILL_PATTERN_LIGHTGRAY),
                                  alignment=Alignment(horizontal='center', vertical='center'),

                                  )
        return header_style

    @classmethod
    def get_table_default_style(cls):
        style = NamedStyle(name='table', font=Font(size=13, bold=False, name='Arial'),
                           alignment=Alignment(horizontal='center', vertical='center'))
        return style

    def create_single_sheet_excel(self, headers_dict: dict, input_data: list, record_type: str = 'json',
                                  sheet_name=None, workbook: openpyxl.Workbook = None, encoding='utf-8'):
        sheet_name = sheet_name or str(datetime.date.today())
        workbook = workbook or self.create_new_workbook(encoding=encoding)
        workbook = self.add_sheet(workbook=workbook, sheet_name=sheet_name)
        data = self.clean_data(headers_dict=headers_dict, input_data=input_data, record_type=record_type)
        workbook = self.write_data_to_sheet(workbook=workbook, sheet_name=sheet_name, data=data)
        header_style = self.get_header_default_style()
        workbook.add_named_style(header_style)
        sheet = workbook.get_sheet_by_name(name=sheet_name)
        for cell in sheet[1]:
            cell.style = header_style
        for col in sheet.columns:
            col_width = 25
            sheet.column_dimensions[col[0].column_letter].width = col_width
        return workbook
